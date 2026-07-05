"""
Equity Research Terminal v4 — Professional Institutional Dashboard
====================================================================
Run:   streamlit run dashboard_v4.py
URL:   http://localhost:8501

Install once:
    pip install streamlit plotly pandas requests beautifulsoup4 pdfplumber yfinance openpyxl feedparser

Keep forensics_engine.py in the same folder.
"""

from st_keyup import st_keyup
import streamlit as st
import pandas as pd
import json, re, os, sqlite3, datetime, io, math

st.set_page_config(
    page_title="Equity Research Terminal",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── Dependency checks ────────────────────────────────────────────────
try:
    import requests
    from bs4 import BeautifulSoup
    SCRAPER_OK = True
except ImportError:
    SCRAPER_OK = False

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    import plotly.express as px
    PLOTLY_OK = True
except ImportError:
    PLOTLY_OK = False

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    import yfinance as yf
    YF_OK = True
except ImportError:
    YF_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XL_OK = True
except ImportError:
    XL_OK = False

try:
    import feedparser
    FEED_OK = True
except ImportError:
    FEED_OK = False

# ─── Forensics engine ────────────────────────────────────────────────
import importlib.util
_fe_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "forensics_engine.py")
if os.path.exists(_fe_path):
    _spec = importlib.util.spec_from_file_location("forensics_engine", _fe_path)
    _fe   = importlib.util.module_from_spec(_spec)
    _spec.loader.exec_module(_fe)
    run_full_forensics = _fe.run_full_forensics
    scan_ar_pdf        = _fe.scan_ar_pdf
    FORENSICS_OK = True
else:
    FORENSICS_OK = False

# ─── Database ────────────────────────────────────────────────────────
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "portfolio.db")

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""CREATE TABLE IF NOT EXISTS portfolio (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ticker TEXT, company TEXT, buy_price REAL,
        qty INTEGER, buy_date TEXT, notes TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS watchlist (
        ticker TEXT PRIMARY KEY, company TEXT, added_date TEXT, notes TEXT)""")
    conn.commit(); conn.close()

init_db()

# ─── CSS ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
body, .stApp { background:#0d0f18 !important; }
.metric-card {
    background:linear-gradient(135deg,#1a1d2e,#1e2235);
    border-radius:10px; padding:14px 18px; margin:4px 0;
    border-left:4px solid #4f8ef7;
    box-shadow:0 2px 8px rgba(0,0,0,0.3);
}
.metric-label{color:#8b92a5;font-size:11px;margin-bottom:3px;text-transform:uppercase;letter-spacing:.5px;}
.metric-value{color:#fff;font-size:20px;font-weight:700;}
.metric-sub{color:#8b92a5;font-size:11px;margin-top:2px;}
.section-title{color:#4f8ef7;font-size:14px;font-weight:700;
    margin:20px 0 10px;padding-bottom:6px;
    border-bottom:2px solid #2d3147;text-transform:uppercase;letter-spacing:.8px;}
.yes-badge{background:#0d2a1a;color:#4caf50;border-radius:6px;
    padding:6px 14px;font-size:13px;margin:2px 0;display:block;
    border-left:3px solid #4caf50;}
.no-badge{background:#2a0d0d;color:#f44336;border-radius:6px;
    padding:6px 14px;font-size:13px;margin:2px 0;display:block;
    border-left:3px solid #f44336;}
.warn-badge{background:#2a1a0d;color:#ff9800;border-radius:6px;
    padding:6px 14px;font-size:13px;margin:2px 0;display:block;
    border-left:3px solid #ff9800;}
.na-badge{background:#1a1a2a;color:#888;border-radius:6px;
    padding:6px 14px;font-size:13px;margin:2px 0;display:block;
    border-left:3px solid #444;}
.data-table{width:100%;border-collapse:collapse;font-size:13px;margin-top:4px;}
.data-table th{
    color:#8b92a5;padding:9px 8px;
    border-bottom:2px solid #2d3147;text-align:right;
    font-weight:600;font-size:11px;text-transform:uppercase;
    background:#141626;position:sticky;top:0;z-index:1;
}
.data-table th:first-child{text-align:left;min-width:180px;}
.data-table td{padding:7px 8px;border-bottom:1px solid #1a1d2e;
    color:#b0b8cc;text-align:right;white-space:nowrap;}
.data-table td:first-child{text-align:left;color:#e0e6f0;min-width:180px;}
.data-table tr.hl td{font-weight:700;color:#fff !important;
    background:linear-gradient(90deg,#1a2035,#161828);}
.data-table tr:hover td{background:#1a1d2e;}
.data-table td.neg{color:#f44336 !important;}
.data-table td.pos{color:#4caf50 !important;}
.news-card{background:#1a1d2e;border-radius:8px;padding:12px 16px;
    margin:6px 0;border-left:3px solid #4f8ef7;
    transition:border-color .2s;}
.news-card:hover{border-left-color:#7ab3ff;}
.news-title{color:#e0e6f0;font-size:14px;font-weight:600;line-height:1.4;}
.news-meta{color:#8b92a5;font-size:11px;margin-top:5px;}
.mgmt-card{background:#1a1d2e;border-radius:10px;padding:14px;
    margin:6px 0;border:1px solid #2d3147;}
.mgmt-name{color:#fff;font-size:14px;font-weight:700;}
.mgmt-title{color:#4f8ef7;font-size:12px;margin-top:2px;}
.mgmt-detail{color:#8b92a5;font-size:11px;margin-top:4px;line-height:1.5;}
.tag{background:#1a2a3a;color:#4f8ef7;padding:3px 8px;
    border-radius:4px;font-size:11px;margin-right:4px;}
.overview-box{background:#1a1d2e;border-radius:10px;padding:16px;
    margin:8px 0;border:1px solid #2d3147;}
[data-testid="stSidebar"]{background:#0d0f18 !important;}
div[data-testid="stTabs"] button{font-size:12px;}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────
# CORE HELPERS
# ─────────────────────────────────────────────────────────────────────

def clean_val(s):
    if s is None: return None
    s = str(s).strip()
    if s in ["-","","—","NA","N/A","nan","-"]: return None
    try:
        return float(s.replace(",","").replace("%","").replace("₹","").replace("Cr","").replace(" ",""))
    except:
        return None

def get_series(rows, label, yr_cols):
    label_l = label.lower().strip()
    for row in rows:
        vals = list(row.values())
        if not vals: continue
        first = str(vals[0]).strip().lower()
        # Match if label is contained in the row label or vice versa
        if label_l in first or first in label_l:
            return {c: clean_val(row.get(c)) for c in yr_cols if c in row}
    return {}

def fmt(val, decimals=0, prefix="", suffix=""):
    if val is None: return "—"
    try:
        f = float(val)
        if decimals == 0:
            return f"{prefix}{f:,.0f}{suffix}"
        return f"{prefix}{f:,.{decimals}f}{suffix}"
    except:
        return str(val)

def render_table(rows, highlight=None, color_negatives=True):
    if not rows:
        return "<p style='color:#555;padding:12px;font-size:13px'>No data available</p>"
    cols = list(rows[0].keys())
    disp = ["Particulars" if c in ("Col0","","0") else c for c in cols]
    hl_lower = [h.lower() for h in (highlight or [])]

    html = "<div style='overflow-x:auto'><table class='data-table'><thead><tr>"
    for dc in disp:
        html += f"<th>{dc}</th>"
    html += "</tr></thead><tbody>"

    for row in rows:
        vals  = list(row.values())
        label = str(vals[0]).strip() if vals else ""
        is_hl = any(h in label.lower() for h in hl_lower)
        html += f"<tr{'  class=\"hl\"' if is_hl else ''}>"
        for i, v in enumerate(vals):
            s = str(v).strip() if v is not None else "—"
            if s == "": s = "—"
            if i == 0:
                html += f"<td>{s}</td>"
            else:
                cls = ""
                if color_negatives:
                    try:
                        num = float(s.replace(",","").replace("%",""))
                        if num < 0: cls = " class='neg'"
                        elif num > 0 and "%" in s and not is_hl: cls = " class='pos'"
                    except: pass
                html += f"<td{cls}>{s}</td>"
        html += "</tr>"
    html += "</tbody></table></div>"
    return html

def bar_chart(years, values, title, color="#4f8ef7", height=260):
    if not PLOTLY_OK: return None
    valid = [(y,v) for y,v in zip(years,values) if v is not None]
    if not valid: return None
    ys, vs = zip(*valid)
    clrs = [color if v >= 0 else "#f44336" for v in vs]
    fig = go.Figure(go.Bar(
        x=list(ys), y=list(vs), marker_color=clrs,
        text=[f"{v:,.0f}" for v in vs],
        textposition="outside", textfont_size=9
    ))
    fig.update_layout(title=dict(text=title, font=dict(size=13)),
        paper_bgcolor="#1a1d2e", plot_bgcolor="#1a1d2e",
        font_color="#b0b8cc", height=height, showlegend=False,
        margin=dict(t=35,b=5,l=5,r=5))
    fig.update_xaxes(showgrid=False, tickfont_size=9, tickangle=-45, linecolor="#2d3147")
    fig.update_yaxes(showgrid=True, gridcolor="#232640", tickfont_size=9, zeroline=True, zerolinecolor="#3d4060")
    return fig

def line_chart(years, series, title, height=270):
    if not PLOTLY_OK: return None
    colors = ["#4f8ef7","#4caf50","#ff9800","#f44336","#9c27b0","#00bcd4","#ff5722"]
    fig = go.Figure()
    for i,(name,vals) in enumerate(series.items()):
        valid_pairs = [(y,v) for y,v in zip(years,vals) if v is not None]
        if not valid_pairs: continue
        ys,vs = zip(*valid_pairs)
        fig.add_trace(go.Scatter(x=list(ys), y=list(vs), name=name, mode="lines+markers",
            line=dict(color=colors[i%len(colors)], width=2),
            marker=dict(size=5, color=colors[i%len(colors)])))
    fig.update_layout(title=dict(text=title, font=dict(size=13)),
        paper_bgcolor="#1a1d2e", plot_bgcolor="#1a1d2e",
        font_color="#b0b8cc", height=height,
        margin=dict(t=35,b=5,l=5,r=5),
        legend=dict(bgcolor="#1a1d2e", bordercolor="#2d3147", font_size=11))
    fig.update_xaxes(showgrid=False, tickfont_size=9, tickangle=-45, linecolor="#2d3147")
    fig.update_yaxes(showgrid=True, gridcolor="#232640", tickfont_size=9)
    return fig

def kpi(col, label, value, unit="", color="#4f8ef7", sub=""):
    col.markdown(
        f"<div class='metric-card' style='border-left-color:{color}'>"
        f"<div class='metric-label'>{label}</div>"
        f"<div class='metric-value'>{value}"
        f"<span style='font-size:11px;color:#8b92a5;margin-left:4px'>{unit}</span></div>"
        f"{'<div class=\"metric-sub\">' + sub + '</div>' if sub else ''}"
        f"</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────
# SCREENER SEARCH AUTOCOMPLETE
# ─────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=300, show_spinner=False)
def search_screener_api(query):
    if not SCRAPER_OK or len(query) < 2: return []
    try:
        url = f"https://www.screener.in/api/company/search/?q={query}&v=3&fts=1"
        r   = requests.get(url,
            headers={"User-Agent":"Mozilla/5.0","X-Requested-With":"XMLHttpRequest"},
            timeout=8)
        if r.status_code == 200:
            out = []
            for item in r.json()[:8]:
                name = item.get("name","")
                if not name: continue
                parts  = [p for p in item.get("url","").strip("/").split("/") if p]
                try:   ticker = parts[parts.index("company") + 1]
                except: ticker = parts[-1] if parts else ""
                clean_name = name.replace(" (Consolidated)","").replace(" (Standalone)","").strip()
                out.append({"name": clean_name, "ticker": ticker})
            return out
    except: pass
    return []

# ─────────────────────────────────────────────────────────────────────
# SCREENER SCRAPER — Full data including all sub-rows
# ─────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner=False)
def scrape_screener(ticker, consolidated=True):
    if not SCRAPER_OK:
        return {"error": "pip install requests beautifulsoup4"}

    suffix = "consolidated" if consolidated else "standalone"
    url    = f"https://www.screener.in/company/{ticker.upper()}/{suffix}/"
    hdrs   = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept-Language": "en-IN,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    try:
        r = requests.get(url, headers=hdrs, timeout=30)
    except Exception as e:
        return {"error": f"Network error: {e}"}

    if r.status_code == 404:
        return {"error": f"'{ticker}' not found. Try Standalone option."}
    if r.status_code != 200:
        return {"error": f"HTTP {r.status_code}"}

    soup   = BeautifulSoup(r.text, "html.parser")
    result = {}

    # Company name
    for sel in ["h1.margin-0","h1"]:
        el = soup.select_one(sel)
        if el:
            result["company_name"] = el.get_text(strip=True)
            break
    else:
        result["company_name"] = ticker

    # Company about/description
    about_el = soup.select_one(".company-profile p") or \
               soup.select_one("#about p") or \
               soup.select_one(".about-section p")
    result["about"] = about_el.get_text(strip=True) if about_el else ""

    # Sector / Industry / BSE / NSE codes
    meta = {}
    for item in soup.select(".company-links a, .company-info li, #company-info li"):
        t = item.get_text(strip=True)
        if t: meta[t] = item.get("href","")
    result["meta"] = meta

    # Market info bar
    info = {}
    for li in soup.select("#top-ratios li"):
        ne = li.select_one("span.name")
        ve = li.select_one("span.number") or li.select_one("span.value")
        if ne and ve:
            info[ne.get_text(strip=True)] = ve.get_text(strip=True)
    result["market_info"] = info

    def clean_cell(td):
        """Extract text — remove expand/collapse buttons but keep label text."""
        # Remove button elements
        for btn in list(td.find_all("button")):
            btn.decompose()
        # Remove icon spans
        for sp in list(td.find_all("span", class_=re.compile(r"icon|arrow|toggle"))):
            sp.decompose()
        text = td.get_text(separator=" ", strip=True)
        # Clean trailing +/- expansion indicators
        text = re.sub(r'\s*[+\-]\s*$', '', text).strip()
        return text

    def parse_section(sec_id):
        """Parse a Screener data table — includes ALL rows (visible + sub-rows)."""
        sec = soup.find("section", {"id": sec_id})
        if not sec: return [], []

        tbl = sec.find("table", class_=re.compile(r"data-table"))
        if not tbl: tbl = sec.find("table")
        if not tbl: return [], []

        # Headers from thead
        thead = tbl.find("thead")
        if thead:
            hcells = thead.find_all("th")
        else:
            all_tr = tbl.find_all("tr")
            hcells = all_tr[0].find_all(["th","td"]) if all_tr else []

        headers = []
        for i, h in enumerate(hcells):
            txt = clean_cell(h).strip()
            headers.append(txt if txt else f"Col{i}")

        if not headers:
            return [], []

        # All body rows — including hidden sub-rows
        tbody   = tbl.find("tbody")
        all_trs = tbody.find_all("tr") if tbody else tbl.find_all("tr")[1:]

        rows_out = []
        for tr in all_trs:
            cells = tr.find_all(["td","th"])
            if not cells: continue

            vals = [clean_cell(c) for c in cells]
            # Pad if fewer cells than headers
            while len(vals) < len(headers):
                vals.append("—")

            # Skip completely empty rows
            if all(v in ["","—","—"," ","-"] for v in vals):
                continue

            # Check if this is a sub-row (indented) — mark it
            is_sub = bool(tr.get("class") and any("sub" in c.lower() for c in tr.get("class",[])))

            row_dict = {headers[i]: vals[i] for i in range(min(len(headers), len(vals)))}

            # For sub-rows, prefix the label with 2 spaces for visual indent
            if is_sub and headers and headers[0] in row_dict:
                row_dict[headers[0]] = "  " + row_dict[headers[0]]

            rows_out.append(row_dict)

        return headers, rows_out

    # Parse all sections
    for sec_id, (ck, rk) in {
        "profit-loss":   ("pnl_cols",   "pnl_rows"),
        "balance-sheet": ("bs_cols",    "bs_rows"),
        "cash-flow":     ("cf_cols",    "cf_rows"),
        "ratios":        ("ratio_cols", "ratio_rows"),
        "quarters":      ("q_cols",     "q_rows"),
        "shareholding":  ("sh_cols",    "sh_rows"),
    }.items():
        c, ro = parse_section(sec_id)
        result[ck] = c
        result[rk] = ro

    # Peers section
    peers = []
    peers_sec = soup.find("section", {"id": "peers"}) or soup.find("section", class_=re.compile(r"peer"))
    if peers_sec:
        for a in peers_sec.find_all("a", href=True)[:10]:
            href = a.get("href","")
            if "/company/" in href:
                pname  = a.get_text(strip=True)
                pticker= [p for p in href.strip("/").split("/") if p]
                if len(pticker) >= 2:
                    peers.append({"name": pname, "ticker": pticker[-1] if pticker[-1] not in ["consolidated","standalone"] else pticker[-2]})
    result["peers"] = peers

    return result


# ─────────────────────────────────────────────────────────────────────
# MANAGEMENT DATA SCRAPER
# ─────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=7200, show_spinner=False)
def scrape_management(ticker, company_name):
    """Scrape management/board details from NSE + MoneyControl."""
    team = []

    if not SCRAPER_OK:
        return team

    hdrs = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    }

    # Try NSE API for company info
    try:
        nse_url = f"https://www.nseindia.com/api/quote-equity?symbol={ticker.upper()}"
        session = requests.Session()
        session.get("https://www.nseindia.com", headers=hdrs, timeout=10)
        r = session.get(nse_url, headers=hdrs, timeout=10)
        if r.status_code == 200:
            data = r.json()
            info = data.get("metadata", {})
            if info.get("pdSectorPe"): pass  # has sector info
    except: pass

    # Try Screener's company page for management info
    try:
        url = f"https://www.screener.in/company/{ticker.upper()}/"
        r   = requests.get(url, headers=hdrs, timeout=15)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, "html.parser")
            # Look for management section
            mgmt_sec = soup.find("section", {"id": "management"}) or \
                       soup.find("div", class_=re.compile(r"management|board|director"))
            if mgmt_sec:
                for item in mgmt_sec.find_all(["li","tr","div"], class_=re.compile(r"person|member|director")):
                    name = item.find(class_=re.compile(r"name"))
                    role = item.find(class_=re.compile(r"role|title|designation"))
                    if name:
                        team.append({
                            "name":  name.get_text(strip=True),
                            "role":  role.get_text(strip=True) if role else "Director",
                            "detail": ""
                        })
    except: pass

    # Try to get annual report link and management discussion from BSE
    try:
        bse_url = f"https://api.bseindia.com/BseIndiaAPI/api/CompanyHeadernew/w?securityCode=&scripCode=&company={company_name[:15]}"
        r = requests.get(bse_url, headers=hdrs, timeout=8)
    except: pass

    # If no management found — create placeholder with useful links
    if not team:
        team = [
            {"name": "Chairman / Managing Director", "role": "See Annual Report for details",
             "detail": f"Download AR from BSE: https://www.bseindia.com | NSE: https://nseindia.com"},
            {"name": "Chief Financial Officer (CFO)", "role": "Key Management Personnel",
             "detail": "Refer to Key Managerial Personnel section in Annual Report"},
            {"name": "Company Secretary", "role": "Compliance Officer",
             "detail": "Contact details available in Annual Report"},
        ]
    return team


# ─────────────────────────────────────────────────────────────────────
# COMPANY-SPECIFIC NEWS (Google News RSS — free, no API key)
# ─────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=600, show_spinner=False)
def get_company_news(ticker, company_name):
    items = []
    if not FEED_OK:
        return _fallback_links(ticker, company_name)

    # Multiple company-specific queries for better coverage
    queries = [
        f'"{company_name}" NSE',
        f'{ticker} NSE stock India',
        f'{company_name} quarterly results',
    ]

    for q in queries[:2]:
        try:
            encoded = requests.utils.quote(q)
            url  = f"https://news.google.com/rss/search?q={encoded}&hl=en-IN&gl=IN&ceid=IN:en"
            feed = feedparser.parse(url)
            for entry in feed.entries[:8]:
                title = entry.get("title","").strip()
                link  = entry.get("link","#")
                pub   = entry.get("published","")
                src   = entry.get("source",{}).get("title","") if hasattr(entry.get("source",{}),"get") else ""
                if title and len(title) > 10:
                    items.append({
                        "title":     title,
                        "link":      link,
                        "published": pub[:16] if pub else "",
                        "source":    src or "Google News"
                    })
            if len(items) >= 8: break
        except: continue

    # De-duplicate
    seen = set()
    unique = []
    for item in items:
        key = item["title"][:50]
        if key not in seen:
            seen.add(key)
            unique.append(item)

    if len(unique) < 3:
        unique += _fallback_links(ticker, company_name)

    return unique[:15]


def _fallback_links(ticker, company_name):
    return [
        {"title": f"NSE — {company_name} Announcements",
         "link":  f"https://www.nseindia.com/get-quotes/equity?symbol={ticker}",
         "published": "Live", "source": "NSE India"},
        {"title": f"BSE — {company_name} Corporate Filings",
         "link":  "https://www.bseindia.com/corporates/ann.html",
         "published": "Live", "source": "BSE India"},
        {"title": f"Screener.in — {company_name} Financial Analysis",
         "link":  f"https://www.screener.in/company/{ticker}/",
         "published": "Live", "source": "Screener.in"},
        {"title": f"Tijori Finance — {company_name} Deep Research",
         "link":  f"https://www.tijorifinance.com/company/{ticker.lower()}/",
         "published": "Live", "source": "Tijori Finance"},
        {"title": f"MoneyControl — {company_name}",
         "link":  f"https://www.moneycontrol.com/",
         "published": "Live", "source": "MoneyControl"},
    ]


# ─────────────────────────────────────────────────────────────────────
# PRICE CHART
# ─────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=900, show_spinner=False)
def get_price_data(ticker, period="2y"):
    if not YF_OK: return None
    for suffix in [".NS",".BO",""]:
        try:
            df = yf.download(ticker+suffix, period=period, auto_adjust=True, progress=False)
            if df is not None and len(df) > 20:
                df.columns = [c[0] if isinstance(c,tuple) else c for c in df.columns]
                return df
        except: continue
    return None


def make_price_chart(df, ticker, show_vol=True):
    if not PLOTLY_OK or df is None: return None
    rows   = 2 if show_vol else 1
    heights= [0.75,0.25] if show_vol else [1.0]
    fig    = make_subplots(rows=rows, cols=1, shared_xaxes=True,
                           vertical_spacing=0.02, row_heights=heights)
    # Candles
    fig.add_trace(go.Candlestick(
        x=df.index, open=df["Open"], high=df["High"],
        low=df["Low"],  close=df["Close"], name="Price",
        increasing_line_color="#4caf50", decreasing_line_color="#f44336",
        increasing_fillcolor="#0d2a1a", decreasing_fillcolor="#2a0d0d"
    ), row=1, col=1)
    # Moving averages
    for d,col,nm in [(20,"#ff9800","MA20"),(50,"#4f8ef7","MA50"),(200,"#9c27b0","MA200")]:
        if len(df) >= d:
            ma = df["Close"].rolling(d).mean()
            fig.add_trace(go.Scatter(x=df.index, y=ma, name=nm,
                line=dict(color=col, width=1.5), opacity=0.9), row=1, col=1)
    # Volume
    if show_vol and "Volume" in df.columns:
        vc = ["#4caf50" if df["Close"].iloc[i] >= df["Open"].iloc[i] else "#f44336"
               for i in range(len(df))]
        fig.add_trace(go.Bar(x=df.index, y=df["Volume"], name="Volume",
            marker_color=vc, opacity=0.6, showlegend=False), row=2, col=1)
    fig.update_layout(
        title=f"{ticker} — Price & Volume",
        paper_bgcolor="#1a1d2e", plot_bgcolor="#1a1d2e",
        font_color="#b0b8cc", height=520,
        margin=dict(t=40,b=10,l=10,r=10),
        xaxis_rangeslider_visible=False,
        legend=dict(bgcolor="#1a1d2e", bordercolor="#2d3147", font_size=11)
    )
    fig.update_xaxes(showgrid=False, tickfont_size=9, linecolor="#2d3147")
    fig.update_yaxes(showgrid=True, gridcolor="#232640", tickfont_size=9)
    return fig


# ─────────────────────────────────────────────────────────────────────
# COMPUTE RATIOS (all 10yr)
# ─────────────────────────────────────────────────────────────────────

def compute_ratios_table(data, yr_cols):
    pnl = data.get("pnl_rows",[])
    bs  = data.get("bs_rows", [])
    cf  = data.get("cf_rows", [])

    sales  = get_series(pnl, "Sales",                 yr_cols)
    net_p  = get_series(pnl, "Net Profit",            yr_cols)
    pbt    = get_series(pnl, "Profit before tax",     yr_cols)
    dep    = get_series(pnl, "Depreciation",          yr_cols)
    intr   = get_series(pnl, "Interest",              yr_cols)
    opm    = get_series(pnl, "OPM",                   yr_cols)
    eps_s  = get_series(pnl, "EPS in Rs",             yr_cols)
    div_po = get_series(pnl, "Dividend Payout",       yr_cols)

    eq_c   = get_series(bs,  "Equity Capital",        yr_cols)
    res_s  = get_series(bs,  "Reserves",              yr_cols)
    borrow = get_series(bs,  "Borrowings",            yr_cols)
    ta_s   = get_series(bs,  "Total Assets",          yr_cols)
    debtor = get_series(bs,  "Debtors",               yr_cols)
    inven  = get_series(bs,  "Inventory",             yr_cols)
    payble = get_series(bs,  "Trade Payables",        yr_cols)
    cash_s = get_series(bs,  "Cash Equivalents",      yr_cols)

    cfo    = get_series(cf,  "Cash from Operating",   yr_cols)
    cfi    = get_series(cf,  "Cash from Investing",   yr_cols)
    capex  = get_series(cf,  "Capital Expenditure",   yr_cols)
    fcf    = get_series(cf,  "Free Cash Flow",        yr_cols)

    metrics = [
        ("── PROFITABILITY ──",   None),
        ("Sales (₹ Cr)",          "sales"),
        ("Operating Profit (Cr)", "op"),
        ("OPM %",                 "opm"),
        ("Net Profit (Cr)",       "np"),
        ("PAT Margin %",          "pat_m"),
        ("PBT Margin %",          "pbt_m"),
        ("EPS (₹)",               "eps"),
        ("Dividend Payout %",     "div_po"),
        ("── RETURNS ──",         None),
        ("ROE %",                 "roe"),
        ("ROCE %",                "roce"),
        ("ROA %",                 "roa"),
        ("── LEVERAGE ──",        None),
        ("Debt / Equity",         "de"),
        ("Interest Coverage",     "ic"),
        ("Net Debt (Cr)",         "net_d"),
        ("── EFFICIENCY ──",      None),
        ("Asset Turnover",        "at"),
        ("Debtor Days",           "dd"),
        ("Inventory Days",        "id"),
        ("Payable Days",          "pd"),
        ("Cash Conversion Cycle", "ccc"),
        ("── CASH FLOW ──",       None),
        ("CFO / PAT",             "cfo_pat"),
        ("CFO / Sales %",         "cfo_s"),
        ("Free Cash Flow (Cr)",   "fcf_v"),
        ("Capex / Sales %",       "cpx_s"),
    ]

    rows = []
    for label, key in metrics:
        if key is None:
            # Section header row
            row = {"Metric": label}
            for yr in yr_cols: row[yr] = ""
            rows.append(row)
            continue

        row = {"Metric": label}
        for yr in yr_cols:
            s  = sales.get(yr);  np = net_p.get(yr); pb = pbt.get(yr)
            dv = dep.get(yr) or 0; iv = intr.get(yr) or 0
            om = opm.get(yr)
            eq = (eq_c.get(yr) or 0) + (res_s.get(yr) or 0)
            ta = ta_s.get(yr);  br = borrow.get(yr) or 0
            cf_= cfo.get(yr);   cx = capex.get(yr)
            db = debtor.get(yr); inv= inven.get(yr); py = payble.get(yr)
            ca = cash_s.get(yr) or 0; fc = fcf.get(yr)
            ep = eps_s.get(yr); dp = div_po.get(yr)

            val = None
            op  = (s - (s * (1 - (om or 0)/100))) if s and om else None
            if s and om: op = round(s * om / 100, 0)

            if   key == "sales"   and s:              val = round(s, 0)
            elif key == "op"      and op:             val = round(op, 0)
            elif key == "opm"     and om:             val = round(om, 1)
            elif key == "np"      and np:             val = round(np, 0)
            elif key == "pat_m"   and np and s and s: val = round(np/s*100, 1)
            elif key == "pbt_m"   and pb and s and s: val = round(pb/s*100, 1)
            elif key == "eps"     and ep:             val = round(ep, 1)
            elif key == "div_po"  and dp:             val = round(dp, 1)
            elif key == "roe"     and np and eq > 0:  val = round(np/eq*100, 1)
            elif key == "roce"    and pb is not None and eq+br > 0:
                val = round((pb+iv)/(eq+br)*100, 1)
            elif key == "roa"     and np and ta and ta > 0: val = round(np/ta*100, 1)
            elif key == "de"      and eq > 0:         val = round(br/eq, 2)
            elif key == "ic"      and iv > 0 and pb:  val = round((pb+iv)/iv, 1)
            elif key == "net_d":                       val = round(br - ca, 0)
            elif key == "at"      and s and ta and ta>0: val = round(s/ta, 2)
            elif key == "dd"      and db and s and s>0: val = round(db/(s/365), 0)
            elif key == "id"      and inv and s and s>0: val = round(inv/(s/365), 0)
            elif key == "pd"      and py and s and s>0: val = round(py/(s/365), 0)
            elif key == "ccc":
                dd_ = round(db/(s/365),0)  if db and s and s>0 else 0
                id_ = round(inv/(s/365),0) if inv and s and s>0 else 0
                pd_ = round(py/(s/365),0)  if py and s and s>0 else 0
                val = round(dd_ + id_ - pd_, 0) if (dd_ or id_ or pd_) else None
            elif key == "cfo_pat" and cf_ and np and np!=0: val = round(cf_/np, 2)
            elif key == "cfo_s"   and cf_ and s and s>0:    val = round(cf_/s*100, 1)
            elif key == "fcf_v"   and fc:             val = round(fc, 0)
            elif key == "cpx_s"   and cx and s and s>0: val = round(abs(cx)/s*100, 1)

            row[yr] = fmt(val) if val is not None else "—"
        rows.append(row)
    return rows


# ─────────────────────────────────────────────────────────────────────
# PEER COMPARISON
# ─────────────────────────────────────────────────────────────────────

def build_peer_table(peer_data):
    companies = list(peer_data.keys())
    rows = []
    sections = [
        ("── VALUATION ──", []),
        ("Market Cap", ("market_info","Market Cap")),
        ("Stock P/E",  ("market_info","Stock P/E")),
        ("Price/Book", ("market_info","Price to book")),
        ("Div Yield %",("market_info","Dividend Yield")),
        ("── FINANCIALS (Latest Year) ──", []),
        ("Revenue (Cr)",    ("pnl","Sales")),
        ("Net Profit (Cr)", ("pnl","Net Profit")),
        ("OPM %",           ("pnl","OPM %")),
        ("EPS (₹)",         ("pnl","EPS in Rs")),
        ("── RETURNS ──", []),
        ("ROCE %",  ("market_info","ROCE")),
        ("ROE %",   ("market_info","ROE")),
        ("── BALANCE SHEET ──", []),
        ("Total Assets",  ("bs","Total Assets")),
        ("Borrowings",    ("bs","Borrowings")),
        ("Equity",        ("bs","Equity Capital")),
    ]

    for label, src in sections:
        if not src:
            row = {"Metric": label}
            for co in companies: row[co] = ""
            rows.append(row); continue

        row = {"Metric": label}
        for co in companies:
            d  = peer_data[co]
            pc = d.get("pnl_cols",[]); yr = [c for c in pc[1:] if c and c!="TTM"]
            kind, key = src

            if kind == "market_info":
                mi = d.get("market_info",{})
                val = "—"
                for k,v in mi.items():
                    if key.lower() in k.lower():
                        val = v; break
                row[co] = val

            elif kind in ("pnl","bs","cf"):
                raw = d.get(f"{'pnl' if kind=='pnl' else kind}_rows",[])
                if yr and raw:
                    s = get_series(raw, key, [yr[-1]])
                    v = s.get(yr[-1])
                    row[co] = fmt(v) if v is not None else "—"
                else:
                    row[co] = "—"
        rows.append(row)
    return rows


# ─────────────────────────────────────────────────────────────────────
# DCF MODEL
# ─────────────────────────────────────────────────────────────────────

def run_dcf(revenue, pat_margin, growth, terminal_growth,
            discount, years, shares, debt, cash):
    pat = revenue * pat_margin / 100
    fcfs, pv_fcfs = [], []
    for i in range(1, years+1):
        projected = pat * ((1 + growth/100) ** i)
        pv        = projected / ((1 + discount/100) ** i)
        fcfs.append(projected)
        pv_fcfs.append(pv)
    tv    = fcfs[-1] * (1 + terminal_growth/100) / (discount/100 - terminal_growth/100)
    pv_tv = tv / ((1 + discount/100) ** years)
    ev    = sum(pv_fcfs) + pv_tv
    eq_v  = ev - debt + cash
    iv    = eq_v / shares if shares > 0 else 0
    return {
        "fcfs": fcfs, "pv_fcfs": pv_fcfs,
        "tv": tv, "pv_tv": pv_tv,
        "ev": ev, "equity_value": eq_v,
        "iv": iv, "years": list(range(1, years+1))
    }


# ─────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────

def export_excel(data, company_name):
    if not XL_OK: return None
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Overview sheet
        minfo = data.get("market_info",{})
        if minfo:
            df_ov = pd.DataFrame(list(minfo.items()), columns=["Metric","Value"])
            df_ov.to_excel(writer, sheet_name="Overview", index=False)

        for sheet, key in [
            ("P&L",          "pnl_rows"),
            ("Balance Sheet","bs_rows"),
            ("Cash Flow",    "cf_rows"),
            ("Ratios",       "ratio_rows"),
            ("Quarterly",    "q_rows"),
            ("Shareholding", "sh_rows"),
        ]:
            rows = data.get(key,[])
            if not rows: continue
            df = pd.DataFrame(rows)
            if "Col0" in df.columns:
                df = df.rename(columns={"Col0":"Particulars"})
            df.to_excel(writer, sheet_name=sheet, index=False)

            ws = writer.sheets[sheet]
            # Format
            ws.column_dimensions["A"].width = 35
            for col_cells in list(ws.columns)[1:]:
                letter = col_cells[0].column_letter
                ws.column_dimensions[letter].width = 12
                for cell in col_cells:
                    cell.alignment = Alignment(horizontal="right")

            # Header style
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1a2035")

    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## 📊 Equity Research")
    st.markdown("---")
    st.markdown("### 🔍 Search Company")
    q = st_keyup(
    "Company name or NSE ticker",
    placeholder="e.g. Solar, EICHERMOT, ITC",
    key="company_search"
)
    sel_ticker = None
    if q and len(q) >= 2 and SCRAPER_OK:
        with st.spinner(""):
            results = search_screener_api(q)
        if results:
            st.markdown("<div style='color:#8b92a5;font-size:11px;margin-bottom:4px'>SUGGESTIONS</div>", unsafe_allow_html=True)
            for i, res in enumerate(results):
                if st.button(f"{res['name']}  ·  {res['ticker']}",
                             key=f"sr_{i}_{res['ticker']}", use_container_width=True):
                    sel_ticker = res["ticker"]
                    st.session_state["active_ticker"] = res["ticker"]
                    st.session_state["active_name"]   = res["name"]
        elif len(q) >= 3:
            st.caption("Not found — try NSE ticker directly")

    st.markdown("---")
    st.markdown("### Direct Ticker Entry")
    ticker_in = st.text_input("NSE Ticker",
                               value=st.session_state.get("active_ticker",""),
                               placeholder="SOLARINDS")
    con_type  = st.radio("Statement", ["Consolidated","Standalone"], horizontal=True)
    fetch_btn = st.button("📥 Fetch 10-Year Data", use_container_width=True)

    st.markdown("---")
    st.markdown("### 📄 Annual Reports (PDF)")
    pdfs = st.file_uploader("Upload PDF(s)", type=["pdf"],
                             accept_multiple_files=True, label_visibility="collapsed")

    st.markdown("---")
    st.markdown("### 📂 Load Saved JSON")
    json_up = st.file_uploader("JSON", type=["json"],
                                accept_multiple_files=True,
                                key="jup", label_visibility="collapsed")

    st.markdown("---")
    st.caption("Built locally · Data: Screener.in + yfinance · Not financial advice")

# ─────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────

if "screener_data" not in st.session_state: st.session_state.screener_data = {}
if "peer_data"     not in st.session_state: st.session_state.peer_data     = {}

if sel_ticker and SCRAPER_OK:
    with st.spinner(f"Loading {sel_ticker}..."):
        d = scrape_screener(sel_ticker, True)
        if "error" not in d:
            st.session_state.screener_data = d
            st.success(f"✓ {d.get('company_name', sel_ticker)}")
        else:
            st.error(d["error"])

if fetch_btn and ticker_in:
    with st.spinner(f"Fetching {ticker_in}..."):
        d = scrape_screener(ticker_in.strip(), con_type == "Consolidated")
        if "error" in d:
            st.error(d["error"])
        else:
            st.session_state.screener_data = d
            st.session_state["active_ticker"] = ticker_in.strip()
            st.success(f"✓ {d.get('company_name', ticker_in)}")

data = st.session_state.screener_data

if json_up and not data:
    try:
        data = json.load(json_up[0])
        st.session_state.screener_data = data
    except: pass

# ─────────────────────────────────────────────────────────────────────
# LANDING PAGE
# ─────────────────────────────────────────────────────────────────────

if not data:
    st.markdown("""
    <div style='text-align:center;padding:40px 0 20px'>
        <div style='font-size:60px'>📊</div>
        <div style='font-size:32px;font-weight:800;color:#fff;margin-top:12px;
             letter-spacing:-0.5px'>Equity Research Terminal</div>
        <div style='color:#8b92a5;font-size:15px;margin-top:8px'>
            Professional · Free · Built for Indian Markets</div>
    </div>""", unsafe_allow_html=True)

    for cols, cards in [
        (st.columns(4), [
            ("🔍","Smart Search","Auto-suggest from Screener.in"),
            ("📈","10yr Financials","P&L, BS, CF with all sub-rows"),
            ("🏥","14 Forensics","Red flags, RPT, Auditor, CARO"),
            ("📉","Price Charts","Candles + MA20/50/200 + RSI"),
        ]),
        (st.columns(4), [
            ("🏆","Peer Compare","Side-by-side vs competitors"),
            ("💹","DCF Model","Intrinsic value + margin of safety"),
            ("💼","Portfolio","Track holdings + watchlist (local)"),
            ("👔","Management","Board & KMP with analysis links"),
        ])
    ]:
        for col,(icon,title,desc) in zip(cols,cards):
            col.markdown(
                f"<div class='metric-card'><div style='font-size:26px'>{icon}</div>"
                f"<div style='color:#fff;font-weight:700;margin-top:8px;font-size:14px'>{title}</div>"
                f"<div style='color:#8b92a5;font-size:12px;margin-top:4px'>{desc}</div></div>",
                unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────────────────────────────
# HEADER + KPIs
# ─────────────────────────────────────────────────────────────────────

company_name  = data.get("company_name","Company")
active_ticker = st.session_state.get("active_ticker", ticker_in or "")
minfo         = data.get("market_info",{})

st.markdown(
    f"<div style='padding:10px 0 4px;display:flex;align-items:center;gap:12px'>"
    f"<span style='font-size:28px;font-weight:800;color:#fff'>{company_name}</span>"
    f"<span style='background:#1a2a3a;color:#4f8ef7;padding:3px 10px;"
    f"border-radius:6px;font-size:12px;font-weight:600'>{active_ticker}</span>"
    f"<span style='color:#8b92a5;font-size:12px'>Screener.in · All values ₹ Crores</span></div>",
    unsafe_allow_html=True)

if minfo:
    mc = st.columns(len(minfo)) if len(minfo) <= 8 else st.columns(8)
    for i,(k,v) in enumerate(list(minfo.items())[:8]):
        mc[i].markdown(
            f"<div class='metric-card' style='padding:10px 14px'>"
            f"<div class='metric-label'>{k}</div>"
            f"<div style='color:#fff;font-size:16px;font-weight:700'>{v}</div></div>",
            unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────
# SHARED DATA
# ─────────────────────────────────────────────────────────────────────

pnl_rows   = data.get("pnl_rows",  [])
bs_rows    = data.get("bs_rows",   [])
cf_rows    = data.get("cf_rows",   [])
q_rows     = data.get("q_rows",    [])
sh_rows    = data.get("sh_rows",   [])
ratio_rows = data.get("ratio_rows",[])
pnl_cols   = data.get("pnl_cols",  [])
bs_cols    = data.get("bs_cols",   [])
cf_cols    = data.get("cf_cols",   [])
q_cols     = data.get("q_cols",    [])
sh_cols    = data.get("sh_cols",   [])
yr_cols    = [c for c in pnl_cols[1:] if c and c != "TTM"] if pnl_cols else []

HL_PNL = ["Sales","Operating Profit","Profit before tax","Net Profit","EBITDA"]
HL_BS  = ["Total Assets","Total Liabilities","Equity Capital","Reserves","Borrowings",
          "Net Worth","Shareholder"]
HL_CF  = ["Cash from Operating","Cash from Investing","Cash from Financing",
          "Net Cash Flow","Free Cash Flow"]
HL_RAT = ["Return on Equity","Return on Capital","Debt to Equity",
          "ROE","ROCE","OPM","Net Profit Margin"]

# ─────────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────────

T = st.tabs([
    "🏠 Overview",
    "📈 P&L",
    "🏦 Balance Sheet",
    "💰 Cash Flow",
    "📊 Ratios",
    "🏥 Forensics",
    "📅 Quarterly",
    "⚖️ Valuation",
    "👥 Shareholding",
    "📉 Price Chart",
    "🏆 Peer Compare",
    "💼 Portfolio",
    "💹 DCF Model",
    "📰 News",
    "👔 Management",
    "📤 Export",
])

# ══ 0. OVERVIEW ══════════════════════════════════════════════════════
with T[0]:
    st.markdown("<div class='section-title'>Company Overview</div>", unsafe_allow_html=True)

    # About
    about = data.get("about","")
    if about:
        st.markdown(f"<div class='overview-box'><div style='color:#b0b8cc;font-size:13px;line-height:1.7'>{about}</div></div>",
                    unsafe_allow_html=True)

    # Quick KPIs
    if minfo and yr_cols and pnl_rows:
        st.markdown("<div class='section-title'>Key Metrics Snapshot</div>", unsafe_allow_html=True)
        ov1,ov2,ov3,ov4,ov5,ov6 = st.columns(6)
        sv = get_series(pnl_rows,"Sales",yr_cols)
        nv = get_series(pnl_rows,"Net Profit",yr_cols)
        ov = get_series(pnl_rows,"OPM %",yr_cols)

        last_s   = list(sv.values())[-1] if sv else None
        last_n   = list(nv.values())[-1] if nv else None
        last_opm = list(ov.values())[-1] if ov else None
        pat_m    = round(last_n/last_s*100,1) if last_s and last_n and last_s>0 else None

        kpi(ov1,"Revenue",fmt(last_s),"Cr","#4f8ef7")
        kpi(ov2,"Net Profit",fmt(last_n),"Cr","#4caf50")
        kpi(ov3,"OPM %",fmt(last_opm,1),"%","#ff9800")
        kpi(ov4,"PAT Margin",fmt(pat_m,1),"%","#9c27b0")

        # Revenue and profit 5yr CAGR
        if len(list(sv.values())) >= 5:
            sv5 = [v for v in sv.values() if v]
            if sv5[0] and sv5[-1] and sv5[0] > 0:
                rev_cagr = round((sv5[-1]/sv5[0])**(1/(len(sv5)-1))*100 - 100, 1)
                kpi(ov5,"Revenue CAGR",f"{rev_cagr:+.1f}","%","#00bcd4",f"{len(sv5)-1}yr")
        nv5 = [v for v in nv.values() if v]
        if len(nv5) >= 5 and nv5[0] and nv5[0] > 0:
            pat_cagr = round((nv5[-1]/nv5[0])**(1/(len(nv5)-1))*100 - 100, 1)
            kpi(ov6,"PAT CAGR",f"{pat_cagr:+.1f}","%","#8bc34a",f"{len(nv5)-1}yr")

    # Charts
    if PLOTLY_OK and yr_cols and pnl_rows:
        st.markdown("<div class='section-title'>Financial Trends</div>", unsafe_allow_html=True)
        sv = get_series(pnl_rows,"Sales",yr_cols)
        nv = get_series(pnl_rows,"Net Profit",yr_cols)
        ov = get_series(pnl_rows,"OPM %",yr_cols)
        c1,c2,c3 = st.columns(3)
        with c1:
            f = bar_chart(list(sv.keys()),list(sv.values()),"Revenue (₹ Cr)","#4f8ef7",240)
            if f: st.plotly_chart(f,use_container_width=True)
        with c2:
            f = bar_chart(list(nv.keys()),list(nv.values()),"Net Profit (₹ Cr)","#4caf50",240)
            if f: st.plotly_chart(f,use_container_width=True)
        with c3:
            f = bar_chart(list(ov.keys()),list(ov.values()),"OPM % Trend","#ff9800",240)
            if f: st.plotly_chart(f,use_container_width=True)

    # Peers from Screener
    peers = data.get("peers",[])
    if peers:
        st.markdown("<div class='section-title'>Sector Peers (from Screener)</div>", unsafe_allow_html=True)
        pc = st.columns(min(5, len(peers)))
        for i,p in enumerate(peers[:5]):
            pc[i%5].markdown(
                f"<div class='metric-card' style='padding:10px 12px;border-left-color:#9c27b0'>"
                f"<div style='color:#fff;font-size:13px;font-weight:600'>{p['name']}</div>"
                f"<div style='color:#9c27b0;font-size:11px'>{p['ticker']}</div></div>",
                unsafe_allow_html=True)


# ══ 1. P&L ═══════════════════════════════════════════════════════════
with T[1]:
    st.markdown("<div class='section-title'>Profit & Loss — 10 Year History</div>", unsafe_allow_html=True)
    if pnl_rows:
        st.markdown(render_table(pnl_rows, HL_PNL), unsafe_allow_html=True)
        if PLOTLY_OK and yr_cols:
            st.markdown("---")
            sv = get_series(pnl_rows,"Sales",yr_cols)
            nv = get_series(pnl_rows,"Net Profit",yr_cols)
            ov = get_series(pnl_rows,"OPM %",yr_cols)
            c1,c2 = st.columns(2)
            with c1:
                f = bar_chart(list(sv.keys()),list(sv.values()),"Revenue (₹ Cr)","#4f8ef7")
                if f: st.plotly_chart(f,use_container_width=True)
            with c2:
                f = bar_chart(list(nv.keys()),list(nv.values()),"Net Profit (₹ Cr)","#4caf50")
                if f: st.plotly_chart(f,use_container_width=True)
            c3,c4 = st.columns(2)
            with c3:
                f = bar_chart(list(ov.keys()),list(ov.values()),"OPM % Trend","#ff9800")
                if f: st.plotly_chart(f,use_container_width=True)
            with c4:
                sv_l = list(sv.values()); yrs = list(sv.keys())
                gr   = [None] + [
                    round((sv_l[i]-sv_l[i-1])/abs(sv_l[i-1])*100,1)
                    if sv_l[i-1] else None for i in range(1,len(sv_l))
                ]
                f = bar_chart(yrs,gr,"Revenue Growth % YoY","#00bcd4")
                if f: st.plotly_chart(f,use_container_width=True)
    else:
        st.info("← Fetch company data from Screener.in using the sidebar")


# ══ 2. BALANCE SHEET ═════════════════════════════════════════════════
with T[2]:
    st.markdown("<div class='section-title'>Balance Sheet — 10 Year History</div>", unsafe_allow_html=True)
    if bs_rows:
        st.markdown(render_table(bs_rows, HL_BS), unsafe_allow_html=True)
        if PLOTLY_OK and yr_cols:
            st.markdown("---")
            eq_s = get_series(bs_rows,"Equity Capital",yr_cols)
            rs_s = get_series(bs_rows,"Reserves",yr_cols)
            br_s = get_series(bs_rows,"Borrowings",yr_cols)
            ta_s = get_series(bs_rows,"Total Assets",yr_cols)
            c1,c2 = st.columns(2)
            with c1:
                nw = [(eq_s.get(y) or 0)+(rs_s.get(y) or 0) for y in yr_cols]
                f  = bar_chart(yr_cols,nw,"Net Worth (Equity + Reserves)","#00bcd4")
                if f: st.plotly_chart(f,use_container_width=True)
            with c2:
                f = bar_chart(list(br_s.keys()),list(br_s.values()),"Borrowings (₹ Cr)","#f44336")
                if f: st.plotly_chart(f,use_container_width=True)
            c3,c4 = st.columns(2)
            with c3:
                f = bar_chart(list(ta_s.keys()),list(ta_s.values()),"Total Assets (₹ Cr)","#9c27b0")
                if f: st.plotly_chart(f,use_container_width=True)
            with c4:
                de = [round((br_s.get(y) or 0)/((eq_s.get(y) or 0)+(rs_s.get(y) or 0)),2)
                      if (eq_s.get(y) or 0)+(rs_s.get(y) or 0)>0 else 0 for y in yr_cols]
                f  = bar_chart(yr_cols,de,"Debt / Equity Ratio","#ff9800")
                if f: st.plotly_chart(f,use_container_width=True)
    else:
        st.info("← Fetch data from Screener.in")


# ══ 3. CASH FLOW ═════════════════════════════════════════════════════
with T[3]:
    st.markdown("<div class='section-title'>Cash Flow — 10 Year History</div>", unsafe_allow_html=True)
    if cf_rows:
        st.markdown(render_table(cf_rows, HL_CF), unsafe_allow_html=True)
        if PLOTLY_OK and yr_cols:
            st.markdown("---")
            cfo_s = get_series(cf_rows,"Cash from Operating",yr_cols)
            cfi_s = get_series(cf_rows,"Cash from Investing",yr_cols)
            cff_s = get_series(cf_rows,"Cash from Financing",yr_cols)
            nv_s  = get_series(pnl_rows,"Net Profit",yr_cols)
            fcf_s = get_series(cf_rows,"Free Cash Flow",yr_cols)
            c1,c2 = st.columns(2)
            with c1:
                f = line_chart(yr_cols,{
                    "Operating CF": list(cfo_s.values()),
                    "Investing CF": list(cfi_s.values()),
                    "Financing CF": list(cff_s.values()),
                },"Cash Flow Trends (₹ Cr)")
                if f: st.plotly_chart(f,use_container_width=True)
            with c2:
                ratio = [round(cfo_s.get(y)/nv_s.get(y),2)
                         if cfo_s.get(y) and nv_s.get(y) and nv_s.get(y)!=0
                         else None for y in yr_cols]
                f = bar_chart(yr_cols,ratio,"CFO / Net Profit (>1 = quality earnings)","#8bc34a")
                if f: st.plotly_chart(f,use_container_width=True)
            c3,c4 = st.columns(2)
            with c3:
                f = bar_chart(list(fcf_s.keys()),list(fcf_s.values()),"Free Cash Flow (₹ Cr)","#4f8ef7")
                if f: st.plotly_chart(f,use_container_width=True)
            with c4:
                sv = get_series(pnl_rows,"Sales",yr_cols)
                cfo_s_pct = [round(cfo_s.get(y)/sv.get(y)*100,1)
                             if cfo_s.get(y) and sv.get(y) and sv.get(y)>0
                             else None for y in yr_cols]
                f = bar_chart(yr_cols,cfo_s_pct,"CFO / Sales % (cash generation)","#ff9800")
                if f: st.plotly_chart(f,use_container_width=True)
    else:
        st.info("← Fetch data from Screener.in")


# ══ 4. RATIOS ════════════════════════════════════════════════════════
with T[4]:
    c_left, c_right = st.columns([2,1])
    with c_left:
        st.markdown("<div class='section-title'>Screener Ratios (10yr)</div>", unsafe_allow_html=True)
        if ratio_rows:
            st.markdown(render_table(ratio_rows, HL_RAT), unsafe_allow_html=True)

    with c_right:
        st.markdown("<div class='section-title'>Latest Ratios</div>", unsafe_allow_html=True)
        if minfo:
            for k,v in minfo.items():
                st.markdown(
                    f"<div style='display:flex;justify-content:space-between;"
                    f"padding:6px 0;border-bottom:1px solid #2d3147;font-size:13px'>"
                    f"<span style='color:#8b92a5'>{k}</span>"
                    f"<span style='color:#fff;font-weight:600'>{v}</span></div>",
                    unsafe_allow_html=True)

    if pnl_rows and yr_cols:
        st.markdown("<div class='section-title'>Computed Ratios — Full 10yr Analysis</div>",
                    unsafe_allow_html=True)
        cr = compute_ratios_table(data, yr_cols)

        # Section headers in different style
        cr_html = "<div style='overflow-x:auto'><table class='data-table'><thead><tr>"
        cols_   = list(cr[0].keys()) if cr else []
        disp_c  = ["Metric" if c == "Metric" else c for c in cols_]
        for dc in disp_c: cr_html += f"<th>{dc}</th>"
        cr_html += "</tr></thead><tbody>"

        for row in cr:
            metric = row.get("Metric","")
            vals   = list(row.values())
            is_sec = metric.startswith("──")
            is_hl  = any(h in metric for h in ["ROE","ROCE","OPM","PAT Margin","Sales (₹","CFO / PAT"])

            if is_sec:
                cr_html += f"<tr><td colspan='{len(vals)}' style='background:#141626;color:#4f8ef7;font-size:11px;font-weight:700;padding:8px;letter-spacing:.8px'>{metric}</td></tr>"
            else:
                cr_html += f"<tr{'  class=\"hl\"' if is_hl else ''}>"
                for i,v in enumerate(vals):
                    if i == 0:
                        cr_html += f"<td>{v}</td>"
                    else:
                        s   = str(v)
                        cls = ""
                        try:
                            f_ = float(s.replace(",",""))
                            if f_ < 0: cls = " class='neg'"
                        except: pass
                        cr_html += f"<td{cls}>{s}</td>"
                cr_html += "</tr>"
        cr_html += "</tbody></table></div>"
        st.markdown(cr_html, unsafe_allow_html=True)

        # Ratio trend charts
        if PLOTLY_OK:
            st.markdown("---")
            def rs(name):
                for row in cr:
                    if row.get("Metric") == name:
                        out = []
                        for y in yr_cols:
                            try: out.append(float(str(row.get(y,"")).replace(",","")))
                            except: out.append(None)
                        return out
                return [None]*len(yr_cols)
            c1,c2,c3 = st.columns(3)
            with c1:
                f = line_chart(yr_cols,{"ROE %":rs("ROE %"),"ROCE %":rs("ROCE %")},"Return Ratios %")
                if f: st.plotly_chart(f,use_container_width=True)
            with c2:
                f = line_chart(yr_cols,{"OPM %":rs("OPM %"),"PAT Margin %":rs("PAT Margin %")},"Margin Trends %")
                if f: st.plotly_chart(f,use_container_width=True)
            with c3:
                f = line_chart(yr_cols,{"Debtor Days":rs("Debtor Days"),"Inventory Days":rs("Inventory Days")},"Working Capital Days")
                if f: st.plotly_chart(f,use_container_width=True)
    else:
        st.info("← Fetch data from Screener.in")


# ══ 5. FORENSICS ═════════════════════════════════════════════════════
with T[5]:
    st.markdown("<div class='section-title'>Forensic Analysis — 14 Point Institutional Checklist</div>",
                unsafe_allow_html=True)

    pdf_findings = None
    with st.expander("📄 Upload Annual Report PDF — auto-checks Auditor Opinion, RPT, Contingent Liabilities, CARO"):
        ar_pdf = st.file_uploader("Scan AR PDF", type=["pdf"], key="for_pdf",
                                   label_visibility="collapsed")
        if ar_pdf and FORENSICS_OK and PDF_OK:
            with st.spinner("Scanning Annual Report..."):
                try:
                    pdf_findings = scan_ar_pdf(ar_pdf.read())
                    if pdf_findings and "error" not in pdf_findings:
                        counts = {
                            "Auditor Flags":  len([x for x in pdf_findings.get("auditor",[])   if x.get("type")=="RED_FLAG"]),
                            "Clean Opinion":  len([x for x in pdf_findings.get("auditor",[])   if x.get("type")=="CLEAN"]),
                            "RPT Alerts":     len([x for x in pdf_findings.get("related_party",[]) if x.get("type")=="ALERT"]),
                            "Contingent L.":  len(pdf_findings.get("contingent_liab",[])),
                            "CARO Flags":     len(pdf_findings.get("caro",[])),
                            "Going Concern":  len(pdf_findings.get("going_concern",[])),
                        }
                        st.success("AR scanned successfully")
                        pc_ = st.columns(6)
                        for i,(k,v) in enumerate(counts.items()):
                            clr = "#f44336" if ("flag" in k.lower() or "alert" in k.lower() or "concern" in k.lower()) and v>0 else "#4caf50"
                            pc_[i].markdown(
                                f"<div class='metric-card' style='border-left-color:{clr};padding:8px 12px'>"
                                f"<div class='metric-label'>{k}</div>"
                                f"<div style='color:{clr};font-size:18px;font-weight:700'>{v}</div></div>",
                                unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Scan error: {e}")
        elif ar_pdf and not PDF_OK:
            st.error("pip install pdfplumber")

    if not pnl_rows:
        st.info("← Fetch data first")
    elif not FORENSICS_OK:
        st.warning("forensics_engine.py not found in same folder — all checks use Screener data only")
        # Run basic forensics inline without the engine
        checks = []
    else:
        checks = run_full_forensics(data, pdf_findings=pdf_findings, ticker=active_ticker)

        yes_n  = sum(1 for c in checks if c["result"]=="yes")
        no_n   = sum(1 for c in checks if c["result"]=="no")
        warn_n = sum(1 for c in checks if c["result"]=="warn")
        na_n   = sum(1 for c in checks if c["result"]=="na")

        s1,s2,s3,s4 = st.columns(4)
        for col,lbl,val,clr in [
            (s1,"✓ PASS",yes_n,"#4caf50"),
            (s2,"✗ FAIL",no_n,"#f44336"),
            (s3,"⚠ WARNING",warn_n,"#ff9800"),
            (s4,"? MANUAL",na_n,"#666"),
        ]:
            col.markdown(
                f"<div class='metric-card' style='border-left-color:{clr};text-align:center'>"
                f"<div class='metric-label'>{lbl}</div>"
                f"<div class='metric-value' style='color:{clr};font-size:28px'>{val}</div></div>",
                unsafe_allow_html=True)

        st.markdown("""
        <div style='margin:12px 0;font-size:11px;color:#8b92a5'>
            <span style='background:#0d1e30;padding:3px 8px;border-radius:4px;margin-right:6px;border:1px solid #1a3a5a'>
                🔵 Screener auto-check</span>
            <span style='background:#0d2a1a;padding:3px 8px;border-radius:4px;margin-right:6px;border:1px solid #1a5a3a'>
                🟢 PDF auto-check</span>
            <span style='background:#2a1a0d;padding:3px 8px;border-radius:4px;border:1px solid #5a3a1a'>
                🟠 Manual verify</span>
        </div>""", unsafe_allow_html=True)

        cats = {}
        for c in checks:
            if c["category"] not in cats: cats[c["category"]] = []
            cats[c["category"]].append(c)

        for cat, cc in cats.items():
            yn = sum(1 for c in cc if c["result"]=="yes")
            nn = sum(1 for c in cc if c["result"]=="no")
            wn = sum(1 for c in cc if c["result"]=="warn")
            hdr_clr = "🔴 " if nn > 0 else "🟡 " if wn > 0 else "🟢 "
            with st.expander(f"{hdr_clr}{cat}  ·  {yn} Pass  {nn} Fail  {wn} Warn",
                             expanded=(nn > 0 or cat.startswith("1."))):
                for c in cc:
                    r    = c["result"]
                    src  = c.get("source","")
                    lnk  = c.get("link","")
                    det  = c.get("detail","")
                    badge= {"yes":"yes-badge","no":"no-badge","warn":"warn-badge"}.get(r,"na-badge")
                    icon = {"yes":"✓","no":"✗","warn":"⚠","na":"?"}.get(r,"?")
                    src_tag = {
                        "screener": "<span style='color:#4f8ef7;font-size:10px;margin-left:8px;background:#0d1e30;padding:2px 6px;border-radius:3px'>SCREENER</span>",
                        "pdf":      "<span style='color:#4caf50;font-size:10px;margin-left:8px;background:#0d2a1a;padding:2px 6px;border-radius:3px'>PDF</span>",
                    }.get(src, "<span style='color:#888;font-size:10px;margin-left:8px;background:#1a1a2a;padding:2px 6px;border-radius:3px'>MANUAL</span>")
                    lnk_tag = f"<a href='{lnk}' target='_blank' style='color:#4f8ef7;font-size:11px;margin-left:6px'>[Open ↗]</a>" if lnk else ""
                    det_html = f"<br><span style='font-size:11px;color:#8b92a5;padding-left:22px;display:block;margin-top:3px'>{det}{lnk_tag}</span>" if det else ""
                    st.markdown(
                        f"<span class='{badge}'>{icon} &nbsp;<b>{c['check']}</b>{src_tag}{det_html}</span>",
                        unsafe_allow_html=True)


# ══ 6. QUARTERLY ═════════════════════════════════════════════════════
with T[6]:
    st.markdown("<div class='section-title'>Quarterly Results</div>", unsafe_allow_html=True)
    if q_rows:
        HL_Q = ["Sales","Net Profit","Operating Profit","Profit before tax"]
        st.markdown(render_table(q_rows, HL_Q), unsafe_allow_html=True)
        if PLOTLY_OK and q_cols:
            qy = [c for c in q_cols[1:] if c]
            qs = get_series(q_rows,"Sales",qy)
            qp = get_series(q_rows,"Net Profit",qy)
            qo = get_series(q_rows,"OPM",qy)
            c1,c2,c3 = st.columns(3)
            with c1:
                f = bar_chart(list(qs.keys()),list(qs.values()),"Quarterly Revenue (₹ Cr)","#4f8ef7")
                if f: st.plotly_chart(f,use_container_width=True)
            with c2:
                f = bar_chart(list(qp.keys()),list(qp.values()),"Quarterly Net Profit (₹ Cr)","#4caf50")
                if f: st.plotly_chart(f,use_container_width=True)
            with c3:
                f = bar_chart(list(qo.keys()),list(qo.values()),"OPM % Quarterly","#ff9800")
                if f: st.plotly_chart(f,use_container_width=True)
    else:
        st.info("← Fetch data from Screener.in")


# ══ 7. VALUATION ═════════════════════════════════════════════════════
with T[7]:
    st.markdown("<div class='section-title'>Valuation Metrics</div>", unsafe_allow_html=True)
    if minfo:
        vc = st.columns(4)
        for i,(k,v) in enumerate(list(minfo.items())[:8]):
            vc[i%4].markdown(
                f"<div class='metric-card'><div class='metric-label'>{k}</div>"
                f"<div class='metric-value'>{v}</div></div>",
                unsafe_allow_html=True)
    if ratio_rows:
        st.markdown("<div class='section-title'>Historical Valuation Ratios</div>",
                    unsafe_allow_html=True)
        val_kw = ["price to earn","price to book","ev/ebitda","dividend yield","market cap",
                  "earning yield","price to sales","price to cash"]
        vr = [r for r in ratio_rows if any(k in str(list(r.values())[0]).lower() for k in val_kw)]
        if vr:
            st.markdown(render_table(vr), unsafe_allow_html=True)
        if PLOTLY_OK and vr and yr_cols:
            pe_s = get_series(vr,"Price to Earning",yr_cols) or get_series(vr,"P/E",yr_cols)
            pb_s = get_series(vr,"Price to Book",yr_cols)
            if pe_s or pb_s:
                series = {}
                if pe_s: series["P/E"] = list(pe_s.values())
                if pb_s: series["P/B"] = list(pb_s.values())
                f = line_chart(yr_cols, series, "Historical Valuation Multiples", 280)
                if f: st.plotly_chart(f, use_container_width=True)
    else:
        st.info("← Fetch data from Screener.in")


# ══ 8. SHAREHOLDING ══════════════════════════════════════════════════
with T[8]:
    st.markdown("<div class='section-title'>Shareholding Pattern</div>", unsafe_allow_html=True)
    if sh_rows:
        st.markdown(render_table(sh_rows), unsafe_allow_html=True)
        if PLOTLY_OK and sh_cols:
            shy = [c for c in sh_cols[1:] if c]
            ps  = get_series(sh_rows,"Promoters",shy)
            fs  = get_series(sh_rows,"FIIs",shy)
            ds  = get_series(sh_rows,"DIIs",shy)
            pbs = get_series(sh_rows,"Public",shy)
            c1,c2 = st.columns(2)
            with c1:
                if ps:
                    f = line_chart(list(ps.keys()),{
                        "Promoters": list(ps.values()),
                        "FII":       list(fs.values()),
                        "DII":       list(ds.values()),
                        "Public":    list(pbs.values()),
                    },"Shareholding Pattern (%)", 300)
                    if f: st.plotly_chart(f,use_container_width=True)
            with c2:
                # Latest shareholding pie
                if ps and PLOTLY_OK:
                    latest_yr = list(ps.keys())[-1]
                    pie_vals  = [ps.get(latest_yr), fs.get(latest_yr), ds.get(latest_yr), pbs.get(latest_yr)]
                    pie_labs  = ["Promoters","FII","DII","Public"]
                    pie_clrs  = ["#4f8ef7","#4caf50","#ff9800","#9c27b0"]
                    pie_data  = [(l,v,c) for l,v,c in zip(pie_labs,pie_vals,pie_clrs) if v]
                    if pie_data:
                        labs,vals,clrs = zip(*pie_data)
                        fig_pie = go.Figure(go.Pie(
                            labels=list(labs), values=list(vals),
                            hole=0.55, marker_colors=list(clrs),
                            textfont_size=12
                        ))
                        fig_pie.update_layout(
                            title=f"Latest: {latest_yr}",
                            paper_bgcolor="#1a1d2e", font_color="#b0b8cc",
                            height=300, margin=dict(t=35,b=5,l=5,r=5),
                            legend=dict(bgcolor="#1a1d2e")
                        )
                        st.plotly_chart(fig_pie, use_container_width=True)
    else:
        st.info("← Fetch data from Screener.in")


# ══ 9. PRICE CHART ═══════════════════════════════════════════════════
with T[9]:
    st.markdown("<div class='section-title'>Price Chart — Technical Analysis</div>",
                unsafe_allow_html=True)
    if not YF_OK:
        st.error("pip install yfinance")
    else:
        pc1,pc2,pc3 = st.columns([2,1,1])
        chart_tk  = pc1.text_input("Ticker", value=active_ticker, placeholder="SOLARINDS", key="ct")
        c_period  = pc2.selectbox("Period", ["1mo","3mo","6mo","1y","2y","3y","5y"], index=4, key="cp")
        show_v    = pc3.checkbox("Volume", value=True, key="cv")

        if chart_tk:
            with st.spinner("Loading price data..."):
                df = get_price_data(chart_tk, c_period)
            if df is not None and len(df) > 0:
                latest = float(df["Close"].iloc[-1])
                high52 = float(df["High"].max())
                low52  = float(df["Low"].min())
                avg_vol= int(df["Volume"].mean()) if "Volume" in df.columns else 0
                chg    = float((df["Close"].iloc[-1]-df["Close"].iloc[-2])/df["Close"].iloc[-2]*100)
                chg1m  = float((df["Close"].iloc[-1]-df["Close"].iloc[-21])/df["Close"].iloc[-21]*100) if len(df)>21 else 0

                k1,k2,k3,k4,k5 = st.columns(5)
                kpi(k1,"LTP",       f"₹{latest:,.2f}",  "",  "#4f8ef7")
                kpi(k2,"1D Change", f"{chg:+.2f}%",     "",  "#4caf50" if chg>=0 else "#f44336")
                kpi(k3,"1M Return", f"{chg1m:+.1f}%",   "",  "#4caf50" if chg1m>=0 else "#f44336")
                kpi(k4,"52W High",  f"₹{high52:,.0f}",  "",  "#4caf50")
                kpi(k5,"52W Low",   f"₹{low52:,.0f}",   "",  "#f44336")

                fig = make_price_chart(df, chart_tk, show_v)
                if fig: st.plotly_chart(fig, use_container_width=True)

                # Indicators
                ind_c1, ind_c2 = st.columns(2)
                with ind_c1:
                    if st.checkbox("RSI (14)", key="rsi_cb"):
                        delta = df["Close"].diff()
                        gain  = delta.clip(lower=0).rolling(14).mean()
                        loss  = (-delta.clip(upper=0)).rolling(14).mean()
                        rsi   = 100 - (100 / (1 + gain/loss.replace(0,1e-10)))
                        fig_r = go.Figure()
                        fig_r.add_trace(go.Scatter(x=df.index, y=rsi,
                            line=dict(color="#ff9800",width=2), name="RSI"))
                        fig_r.add_hline(y=70, line_dash="dash", line_color="#f44336", annotation_text="OB 70")
                        fig_r.add_hline(y=30, line_dash="dash", line_color="#4caf50", annotation_text="OS 30")
                        fig_r.add_hrect(y0=30, y1=70, fillcolor="#4f8ef7", opacity=0.05)
                        fig_r.update_layout(title="RSI (14)", paper_bgcolor="#1a1d2e",
                            plot_bgcolor="#1a1d2e", font_color="#b0b8cc",
                            height=200, margin=dict(t=35,b=5,l=5,r=5), showlegend=False)
                        fig_r.update_xaxes(showgrid=False, tickfont_size=9)
                        fig_r.update_yaxes(showgrid=True, gridcolor="#232640", range=[0,100])
                        st.plotly_chart(fig_r, use_container_width=True)

                with ind_c2:
                    if st.checkbox("MACD (12,26,9)", key="macd_cb"):
                        ema12 = df["Close"].ewm(span=12).mean()
                        ema26 = df["Close"].ewm(span=26).mean()
                        macd  = ema12 - ema26
                        signal= macd.ewm(span=9).mean()
                        hist  = macd - signal
                        fig_m = go.Figure()
                        fig_m.add_trace(go.Bar(x=df.index, y=hist,
                            marker_color=["#4caf50" if v >= 0 else "#f44336" for v in hist],
                            name="Histogram", opacity=0.7))
                        fig_m.add_trace(go.Scatter(x=df.index, y=macd,
                            line=dict(color="#4f8ef7",width=1.5), name="MACD"))
                        fig_m.add_trace(go.Scatter(x=df.index, y=signal,
                            line=dict(color="#ff9800",width=1.5), name="Signal"))
                        fig_m.update_layout(title="MACD (12,26,9)", paper_bgcolor="#1a1d2e",
                            plot_bgcolor="#1a1d2e", font_color="#b0b8cc",
                            height=200, margin=dict(t=35,b=5,l=5,r=5),
                            legend=dict(bgcolor="#1a1d2e", font_size=10))
                        fig_m.update_xaxes(showgrid=False, tickfont_size=9)
                        fig_m.update_yaxes(showgrid=True, gridcolor="#232640")
                        st.plotly_chart(fig_m, use_container_width=True)
            else:
                st.warning(f"No price data for '{chart_tk}'. Make sure it's a valid NSE ticker.")


# ══ 10. PEER COMPARE ═════════════════════════════════════════════════
with T[10]:
    st.markdown("<div class='section-title'>Peer Comparison</div>", unsafe_allow_html=True)

    peer_data = st.session_state.peer_data
    if data: peer_data[company_name] = data

    peer_input = st.text_input("Enter peer NSE tickers (comma-separated, max 5)",
                                placeholder="e.g. BAJAJ-AUTO, TVSMOTORS, HEROMOTOCO")
    p_type     = st.radio("Statement", ["Consolidated","Standalone"],
                           horizontal=True, key="pt")
    load_btn   = st.button("📊 Load Peer Data", key="lp")

    if load_btn and peer_input:
        tickers_ = [t.strip().upper() for t in peer_input.split(",") if t.strip()][:5]
        prog_    = st.progress(0)
        for i, tk in enumerate(tickers_):
            with st.spinner(f"Loading {tk}..."):
                d = scrape_screener(tk, p_type == "Consolidated")
                if "error" not in d:
                    peer_data[d.get("company_name",tk)] = d
                else:
                    st.warning(f"{tk}: {d['error']}")
            prog_.progress((i+1)/len(tickers_))
        st.session_state.peer_data = peer_data
        prog_.empty()
        st.success(f"✓ Loaded {len(peer_data)} companies for comparison")

    if len(peer_data) >= 1:
        ptable = build_peer_table(peer_data)
        # Style section headers
        peer_html = "<div style='overflow-x:auto'><table class='data-table'><thead><tr>"
        peer_cols = list(ptable[0].keys()) if ptable else []
        for pc_ in peer_cols: peer_html += f"<th>{pc_}</th>"
        peer_html += "</tr></thead><tbody>"
        for row in ptable:
            metric = row.get("Metric","")
            vals   = list(row.values())
            is_sec = metric.startswith("──")
            if is_sec:
                peer_html += f"<tr><td colspan='{len(vals)}' style='background:#141626;color:#4f8ef7;font-size:11px;font-weight:700;padding:8px;letter-spacing:.8px'>{metric}</td></tr>"
            else:
                peer_html += "<tr>"
                for i,v in enumerate(vals):
                    peer_html += f"<td>{'<b>' if i==0 else ''}{v}{'</b>' if i==0 else ''}</td>"
                peer_html += "</tr>"
        peer_html += "</tbody></table></div>"
        st.markdown(peer_html, unsafe_allow_html=True)

        # Visual charts
        if PLOTLY_OK and len(peer_data) >= 2:
            st.markdown("---")
            companies_ = list(peer_data.keys())
            for metric_, pnl_key_, color_ in [
                ("Revenue — Latest Year (₹ Cr)","Sales","#4f8ef7"),
                ("Net Profit — Latest Year (₹ Cr)","Net Profit","#4caf50"),
                ("OPM % — Latest Year","OPM %","#ff9800"),
            ]:
                vals_ = []
                for co in companies_:
                    d_ = peer_data[co]
                    pr_ = d_.get("pnl_rows",[]); pc_ = d_.get("pnl_cols",[])
                    yr_ = [c for c in pc_[1:] if c and c != "TTM"]
                    if yr_ and pr_:
                        s_ = get_series(pr_, pnl_key_, [yr_[-1]])
                        vals_.append(s_.get(yr_[-1]))
                    else: vals_.append(None)
                fig_ = go.Figure(go.Bar(
                    x=companies_, y=vals_, marker_color=color_,
                    text=[f"{v:,.0f}" if v else "—" for v in vals_],
                    textposition="outside", textfont_size=10
                ))
                fig_.update_layout(title=metric_, paper_bgcolor="#1a1d2e",
                    plot_bgcolor="#1a1d2e", font_color="#b0b8cc", height=260,
                    showlegend=False, margin=dict(t=35,b=20,l=5,r=5))
                fig_.update_xaxes(showgrid=False, tickfont_size=10)
                fig_.update_yaxes(showgrid=True, gridcolor="#232640")
                st.plotly_chart(fig_, use_container_width=True)
    else:
        st.info("Enter peer tickers above and click Load Peer Data")


# ══ 11. PORTFOLIO ════════════════════════════════════════════════════
with T[11]:
    st.markdown("<div class='section-title'>Portfolio Tracker</div>", unsafe_allow_html=True)
    po1, po2 = st.columns([1.4, 2])

    with po1:
        st.markdown("**Add Holding**")
        pt_  = st.text_input("NSE Ticker", placeholder="SOLARINDS", key="pt_")
        pco_ = st.text_input("Company",    placeholder="Solar Industries", key="pco_")
        ppr_ = st.number_input("Buy Price (₹)", min_value=0.0, step=1.0, key="ppr_")
        pqt_ = st.number_input("Quantity",  min_value=1, step=1, key="pqt_")
        pdt_ = st.date_input("Buy Date",    key="pdt_")
        pnt_ = st.text_input("Notes",       placeholder="Thesis...", key="pnt_")
        if st.button("➕ Add Holding", use_container_width=True, key="padd"):
            if pt_ and ppr_ > 0 and pqt_ > 0:
                conn_ = sqlite3.connect(DB_PATH)
                conn_.execute("INSERT INTO portfolio (ticker,company,buy_price,qty,buy_date,notes) VALUES (?,?,?,?,?,?)",
                              (pt_.upper(), pco_, ppr_, pqt_, str(pdt_), pnt_))
                conn_.commit(); conn_.close()
                st.success("✓ Added"); st.rerun()
            else:
                st.error("Fill ticker, price, qty")

    with po2:
        st.markdown("**My Holdings**")
        conn_  = sqlite3.connect(DB_PATH)
        holds_ = conn_.execute("SELECT * FROM portfolio ORDER BY id DESC").fetchall()
        conn_.close()

        if holds_:
            port_rows_ = []
            tot_inv = tot_cur = 0
            for row_ in holds_:
                id_,tk_,co_,bp_,qty_,bd_,nt_ = row_
                cur_ = bp_
                if YF_OK:
                    try:
                        dfp = yf.download(tk_+".NS", period="2d", auto_adjust=True, progress=False)
                        if dfp is not None and len(dfp) > 0:
                            dfp.columns = [c[0] if isinstance(c,tuple) else c for c in dfp.columns]
                            cur_ = float(dfp["Close"].iloc[-1])
                    except: pass
                inv_  = bp_ * qty_
                curr_ = cur_ * qty_
                pnl_  = curr_ - inv_
                pct_  = pnl_ / inv_ * 100 if inv_ > 0 else 0
                tot_inv += inv_; tot_cur += curr_
                port_rows_.append({
                    "ID": id_, "Ticker": tk_, "Company": co_ or tk_,
                    "Buy ₹": f"{bp_:,.2f}", "Qty": qty_,
                    "CMP ₹": f"{cur_:,.2f}",
                    "Invested": f"{inv_:,.0f}",
                    "Value": f"{curr_:,.0f}",
                    "P&L": f"{pnl_:+,.0f}",
                    "P&L %": f"{pct_:+.1f}%",
                })

            tot_pnl = tot_cur - tot_inv
            pnl_pct = tot_pnl/tot_inv*100 if tot_inv > 0 else 0
            k1_,k2_,k3_ = st.columns(3)
            kpi(k1_,"Total Invested",   f"₹{tot_inv:,.0f}","","#4f8ef7")
            kpi(k2_,"Current Value",    f"₹{tot_cur:,.0f}","","#9c27b0")
            kpi(k3_,"Total P&L",        f"₹{tot_pnl:+,.0f}","",
                "#4caf50" if tot_pnl>=0 else "#f44336", f"{pnl_pct:+.1f}%")

            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown(render_table(port_rows_,["P&L %"]), unsafe_allow_html=True)

            del_id_ = st.number_input("Delete by ID", min_value=0, step=1, key="did_")
            if st.button("🗑️ Delete", key="pdel") and del_id_ > 0:
                conn_ = sqlite3.connect(DB_PATH)
                conn_.execute("DELETE FROM portfolio WHERE id=?", (int(del_id_),))
                conn_.commit(); conn_.close(); st.rerun()
        else:
            st.info("No holdings yet. Add your first stock on the left.")

    # Watchlist
    st.markdown("---")
    st.markdown("<div class='section-title'>Watchlist</div>", unsafe_allow_html=True)
    w1_, w2_ = st.columns([1.2, 2])
    with w1_:
        wtk_ = st.text_input("Ticker",  placeholder="HDFCBANK", key="wtk_")
        wco_ = st.text_input("Company", key="wco_")
        wnt_ = st.text_input("Why watching / Thesis", key="wnt_")
        if st.button("👁️ Add to Watchlist", use_container_width=True, key="wadd"):
            if wtk_:
                conn_ = sqlite3.connect(DB_PATH)
                conn_.execute("INSERT OR REPLACE INTO watchlist (ticker,company,added_date,notes) VALUES (?,?,?,?)",
                              (wtk_.upper(), wco_, str(datetime.date.today()), wnt_))
                conn_.commit(); conn_.close()
                st.success("✓ Added"); st.rerun()
    with w2_:
        conn_  = sqlite3.connect(DB_PATH)
        wrows_ = conn_.execute("SELECT ticker,company,added_date,notes FROM watchlist").fetchall()
        conn_.close()
        if wrows_:
            wdata_ = [{"Ticker":r[0],"Company":r[1] or r[0],"Added":r[2],"Thesis":r[3] or ""} for r in wrows_]
            st.markdown(render_table(wdata_), unsafe_allow_html=True)
            wdel_ = st.text_input("Remove ticker", key="wdel_")
            if st.button("Remove", key="wdelb") and wdel_:
                conn_ = sqlite3.connect(DB_PATH)
                conn_.execute("DELETE FROM watchlist WHERE ticker=?", (wdel_.upper(),))
                conn_.commit(); conn_.close(); st.rerun()
        else:
            st.info("Watchlist is empty")


# ══ 12. DCF MODEL ════════════════════════════════════════════════════
with T[12]:
    st.markdown("<div class='section-title'>DCF — Intrinsic Value Calculator</div>",
                unsafe_allow_html=True)

    # Auto-fill from loaded data
    auto_rev = auto_pat = auto_dep = auto_borrow = auto_cash = 0.0
    auto_shares = 1.0
    if pnl_rows and yr_cols:
        sv_ = get_series(pnl_rows,"Sales",yr_cols)
        nv_ = get_series(pnl_rows,"Net Profit",yr_cols)
        last_s_ = list(sv_.values())[-1] if sv_ else 0
        last_n_ = list(nv_.values())[-1] if nv_ else 0
        auto_rev = float(last_s_ or 0)
        auto_pat = round(float(last_n_/last_s_*100),1) if last_s_ and last_n_ else 10.0
    if bs_rows and yr_cols:
        br_ = get_series(bs_rows,"Borrowings",yr_cols)
        ca_ = get_series(bs_rows,"Cash Equivalents",yr_cols)
        eq_ = get_series(bs_rows,"Equity Capital",yr_cols)
        auto_borrow = float(list(br_.values())[-1] or 0) if br_ else 0
        auto_cash   = float(list(ca_.values())[-1] or 0) if ca_ else 0
        eq_last     = float(list(eq_.values())[-1] or 0) if eq_ else 1
        if minfo.get("Market Cap"):
            try:
                mc_str = minfo["Market Cap"].replace(",","").replace("Cr","").strip()
                mc_val = float(mc_str)
                auto_shares = max(0.01, round(eq_last / 10, 2)) if eq_last else 1.0
            except: pass

    st.markdown("**Inputs — auto-filled from latest data, adjust for your scenario**")
    dc1,dc2,dc3 = st.columns(3)
    rev_d      = dc1.number_input("Latest Revenue (₹ Cr)",      value=auto_rev,    step=100.0)
    pat_d      = dc1.number_input("PAT Margin %",                value=auto_pat,    step=0.5, min_value=0.0, max_value=100.0)
    growth_d   = dc2.number_input("Revenue Growth % (projected)",value=15.0,        step=1.0)
    tg_d       = dc2.number_input("Terminal Growth %",           value=5.0,         step=0.5, min_value=1.0, max_value=8.0)
    disc_d     = dc3.number_input("Discount Rate / WACC %",      value=12.0,        step=0.5)
    yrs_d      = dc3.number_input("Projection Years",            value=10,          step=1, min_value=3, max_value=15)
    dc4,dc5,dc6= st.columns(3)
    shr_d      = dc4.number_input("Shares Outstanding (Cr)",     value=auto_shares, step=0.1, min_value=0.01)
    dbt_d      = dc5.number_input("Total Debt (₹ Cr)",           value=auto_borrow, step=10.0)
    csh_d      = dc6.number_input("Cash & Equivalents (₹ Cr)",   value=auto_cash,   step=10.0)
    cmp_d      = st.number_input("Current Market Price (₹) — for margin of safety", value=0.0, step=10.0)

    c_bear, c_base, c_bull = st.columns(3)
    scenarios = [
        ("🐻 Bear Case", growth_d*0.6, disc_d*1.1, c_bear),
        ("📊 Base Case", growth_d,     disc_d,      c_base),
        ("🐂 Bull Case", growth_d*1.4, disc_d*0.9,  c_bull),
    ]

    if st.button("🧮 Calculate All Scenarios", use_container_width=True, key="dcf_calc"):
        for label, g, d, col in scenarios:
            if d - tg_d <= 0:
                col.warning("Invalid: Discount rate must be > Terminal growth rate")
                continue
            res = run_dcf(rev_d, pat_d, g, tg_d, d, int(yrs_d), shr_d, dbt_d, csh_d)
            clr = {"🐻 Bear Case":"#f44336","📊 Base Case":"#4f8ef7","🐂 Bull Case":"#4caf50"}[label]
            col.markdown(
                f"<div class='metric-card' style='border-left-color:{clr};text-align:center'>"
                f"<div style='color:{clr};font-size:16px;font-weight:700;margin-bottom:8px'>{label}</div>"
                f"<div style='color:#8b92a5;font-size:11px'>Growth: {g:.1f}% | Disc: {d:.1f}%</div>"
                f"<div style='color:#fff;font-size:24px;font-weight:800;margin-top:8px'>"
                f"₹{res['iv']:,.0f}</div>"
                f"<div style='color:#8b92a5;font-size:11px'>Intrinsic Value / Share</div>"
                + (f"<div style='color:{clr};font-size:14px;font-weight:600;margin-top:6px'>"
                   f"MoS: {(res['iv']-cmp_d)/res['iv']*100:+.1f}%</div>" if cmp_d > 0 else "")
                + "</div>",
                unsafe_allow_html=True)

        # FCF chart for base case
        if PLOTLY_OK:
            base = run_dcf(rev_d, pat_d, growth_d, tg_d, disc_d, int(yrs_d), shr_d, dbt_d, csh_d)
            fig_d = go.Figure()
            fig_d.add_trace(go.Bar(x=[f"Yr {y}" for y in base["years"]],
                y=base["fcfs"], name="Projected FCF",
                marker_color="#4f8ef7",
                text=[f"₹{v:,.0f}" for v in base["fcfs"]], textposition="outside", textfont_size=9))
            fig_d.update_layout(title="Base Case — Projected Free Cash Flows (₹ Cr)",
                paper_bgcolor="#1a1d2e", plot_bgcolor="#1a1d2e", font_color="#b0b8cc",
                height=280, showlegend=False, margin=dict(t=35,b=5,l=5,r=5))
            fig_d.update_xaxes(showgrid=False); fig_d.update_yaxes(showgrid=True, gridcolor="#232640")
            st.plotly_chart(fig_d, use_container_width=True)

        st.markdown("""<div style='padding:10px;background:#1a1d2e;border-radius:8px;
            font-size:12px;color:#8b92a5;margin-top:8px'>
            ⚠️ DCF is highly sensitive to assumptions. Use all 3 scenarios to build a range.
            A stock trading below the Bear Case intrinsic value generally has a large margin of safety.
            This is a starting point for analysis — not a buy/sell recommendation.
            </div>""", unsafe_allow_html=True)


# ══ 13. NEWS ═════════════════════════════════════════════════════════
with T[13]:
    st.markdown(f"<div class='section-title'>News & Announcements — {company_name}</div>",
                unsafe_allow_html=True)

    if active_ticker or company_name:
        ticker_for_news = active_ticker or company_name
        co_for_news     = company_name or active_ticker

        with st.spinner(f"Loading news for {co_for_news}..."):
            news_items = get_company_news(ticker_for_news, co_for_news)

        if news_items:
            st.markdown(f"<div style='color:#8b92a5;font-size:12px;margin-bottom:10px'>"
                        f"Showing latest news for <b style='color:#4f8ef7'>{co_for_news}</b> "
                        f"({len(news_items)} articles)</div>", unsafe_allow_html=True)
            for item in news_items:
                pub = item.get("published","")[:16]
                src = item.get("source","")
                st.markdown(
                    f"<div class='news-card'>"
                    f"<div class='news-title'>"
                    f"<a href='{item['link']}' target='_blank' style='color:#e0e6f0;text-decoration:none'>"
                    f"{item['title']}</a></div>"
                    f"<div class='news-meta'>📰 {src}"
                    f"{'  ·  ' + pub if pub else ''}</div>"
                    f"</div>", unsafe_allow_html=True)

        # Direct research links
        st.markdown("---")
        st.markdown("<div class='section-title'>Quick Research Links</div>", unsafe_allow_html=True)
        links_ = [
            ("NSE Announcements",      f"https://www.nseindia.com/get-quotes/equity?symbol={active_ticker}","#4f8ef7"),
            ("BSE Corporate Filings",  "https://www.bseindia.com/corporates/ann.html","#ff9800"),
            ("Screener.in Analysis",   f"https://www.screener.in/company/{active_ticker}/","#4caf50"),
            ("Tijori Finance",         f"https://www.tijorifinance.com/company/{active_ticker.lower()}/","#9c27b0"),
            ("MoneyControl",           f"https://www.moneycontrol.com/","#f44336"),
            ("Economic Times Markets", "https://economictimes.indiatimes.com/markets/stocks","#00bcd4"),
            ("NSE Annual Reports",     f"https://www.nseindia.com/companies-listing/corporate-filings-annual-reports","#ff5722"),
            ("BSE Annual Reports",     "https://www.bseindia.com/stock-share-price/","#8bc34a"),
        ]
        lc_ = st.columns(4)
        for i,(name_,url_,clr_) in enumerate(links_):
            lc_[i%4].markdown(
                f"<a href='{url_}' target='_blank' style='text-decoration:none'>"
                f"<div class='metric-card' style='border-left-color:{clr_};cursor:pointer'>"
                f"<div style='color:{clr_};font-size:13px;font-weight:600'>🔗 {name_}</div>"
                f"</div></a>", unsafe_allow_html=True)
    else:
        st.info("← Load a company first to see company-specific news")


# ══ 14. MANAGEMENT ═══════════════════════════════════════════════════
with T[14]:
    st.markdown(f"<div class='section-title'>Management & Board — {company_name}</div>",
                unsafe_allow_html=True)

    # Try to load management data
    if active_ticker or company_name:
        with st.spinner("Loading management information..."):
            mgmt = scrape_management(active_ticker or "", company_name)

        if mgmt:
            mc_ = st.columns(2)
            for i, person in enumerate(mgmt):
                mc_[i%2].markdown(
                    f"<div class='mgmt-card'>"
                    f"<div class='mgmt-name'>👤 {person['name']}</div>"
                    f"<div class='mgmt-title'>{person['role']}</div>"
                    f"{'<div class=\"mgmt-detail\">' + person['detail'] + '</div>' if person.get('detail') else ''}"
                    f"</div>", unsafe_allow_html=True)

    # Key Management Personnel categories
    st.markdown("---")
    st.markdown("<div class='section-title'>Where to Find Detailed Management Info (Free)</div>",
                unsafe_allow_html=True)

    sources_ = [
        ("Annual Report (AR)",
         f"BSE Filings: https://www.bseindia.com",
         "Director's Report, Management Discussion & Analysis (MDA), Corporate Governance Report — all have detailed management info"),
        ("NSE Company Info",
         f"https://www.nseindia.com/get-quotes/equity?symbol={active_ticker}",
         "Board of Directors, Key Managerial Personnel, registered office details"),
        ("MCA (Ministry of Corporate Affairs)",
         "https://www.mca.gov.in/",
         "Company master data, director DIN details, all official filings"),
        ("Screener.in",
         f"https://www.screener.in/company/{active_ticker}/",
         "Management Q&A, concall transcripts, management tone analysis"),
        ("Tijori Finance",
         f"https://www.tijorifinance.com/company/{active_ticker.lower()}/",
         "Operational metrics, management commentary from investor presentations"),
        ("Trendlyne",
         f"https://trendlyne.com/equity/{active_ticker}//",
         "Management scorecard, pledging details, insider transactions"),
    ]

    for title_, url_, desc_ in sources_:
        st.markdown(
            f"<div class='overview-box'>"
            f"<div style='color:#fff;font-size:14px;font-weight:700'>{title_}</div>"
            f"<a href='{url_}' target='_blank' style='color:#4f8ef7;font-size:12px'>{url_}</a>"
            f"<div style='color:#8b92a5;font-size:12px;margin-top:6px'>{desc_}</div>"
            f"</div>", unsafe_allow_html=True)

    # What to look for in management analysis
    st.markdown("---")
    st.markdown("<div class='section-title'>Management Analysis Checklist</div>",
                unsafe_allow_html=True)
    checklist_ = [
        ("Promoter Holding %",        "Is it >50%? Increasing or decreasing over last 8 quarters?"),
        ("Promoter Pledge %",         "Should be <5%. High pledge = financial stress signal."),
        ("Salary to PAT ratio",       "MD/CEO salary should be <5% of PAT for good governance"),
        ("Related Party Transactions","Are loans/advances given to promoter entities? Check AR notes"),
        ("Capital Allocation",        "Capex > Depreciation? Where is free cash deployed (dividends, buyback, debt paydown)?"),
        ("Guidance Track Record",     "Compare management guidance from past concalls vs actual results"),
        ("Auditor tenure",            "Same Big-4 auditor for >10yrs with no qualification = good sign"),
        ("Board Independence",        "% of independent directors. Do they have relevant expertise?"),
        ("Insider Transactions",      "Any recent insider buying/selling? Check NSE bulk deals"),
        ("Contingent Liabilities",    "Large undisclosed liabilities vs PAT can be ticking time bombs"),
    ]
    cl_ = st.columns(2)
    for i,(point_,detail_) in enumerate(checklist_):
        cl_[i%2].markdown(
            f"<div class='mgmt-card'>"
            f"<div class='mgmt-name'>📋 {point_}</div>"
            f"<div class='mgmt-detail'>{detail_}</div>"
            f"</div>", unsafe_allow_html=True)


# ══ 15. EXPORT ═══════════════════════════════════════════════════════
with T[15]:
    st.markdown("<div class='section-title'>Export Research Report</div>", unsafe_allow_html=True)
    ec1, ec2 = st.columns(2)

    with ec1:
        st.markdown("**Full Excel Report**")
        st.markdown("""
        <div class='overview-box' style='font-size:13px;color:#8b92a5'>
            Exports all tabs: Overview, P&L, Balance Sheet, Cash Flow,
            Ratios, Quarterly Results, Shareholding — formatted for analysis
        </div>""", unsafe_allow_html=True)
        if XL_OK:
            if st.button("📊 Generate Excel", use_container_width=True, key="gen_xl"):
                with st.spinner("Generating Excel..."):
                    buf_ = export_excel(data, company_name)
                if buf_:
                    fname_ = f"{company_name.replace(' ','_')}_Research_{datetime.date.today()}.xlsx"
                    st.download_button("⬇️ Download Excel", data=buf_, file_name=fname_,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_xl")
        else:
            st.error("pip install openpyxl")

    with ec2:
        st.markdown("**Save as JSON**")
        st.markdown("""
        <div class='overview-box' style='font-size:13px;color:#8b92a5'>
            Save all scraped data to JSON — reload later without re-fetching.
            Use with Load Saved JSON in sidebar.
        </div>""", unsafe_allow_html=True)
        if st.button("💾 Save JSON", use_container_width=True, key="gen_json"):
            json_str_ = json.dumps(data, indent=2, default=str)
            fname_j   = f"{company_name.replace(' ','_')}_{datetime.date.today()}.json"
            st.download_button("⬇️ Download JSON", data=json_str_, file_name=fname_j,
                mime="application/json", key="dl_json")

    # Quick summary snapshot
    st.markdown("---")
    st.markdown("<div class='section-title'>Research Summary Snapshot</div>", unsafe_allow_html=True)
    if data:
        ss1,ss2,ss3 = st.columns(3)
        with ss1:
            st.markdown("**Market Info**")
            for k_,v_ in list(minfo.items())[:6]:
                st.markdown(
                    f"<div style='display:flex;justify-content:space-between;"
                    f"padding:5px 0;border-bottom:1px solid #2d3147;font-size:12px'>"
                    f"<span style='color:#8b92a5'>{k_}</span>"
                    f"<span style='color:#fff;font-weight:600'>{v_}</span></div>",
                    unsafe_allow_html=True)
        with ss2:
            st.markdown("**Latest P&L**")
            if pnl_rows and yr_cols:
                for lbl_,key_ in [("Revenue","Sales"),("Net Profit","Net Profit"),
                                   ("OPM %","OPM %"),("EPS","EPS in Rs")]:
                    s_ = get_series(pnl_rows,key_,[yr_cols[-1]])
                    v_ = s_.get(yr_cols[-1])
                    st.markdown(
                        f"<div style='display:flex;justify-content:space-between;"
                        f"padding:5px 0;border-bottom:1px solid #2d3147;font-size:12px'>"
                        f"<span style='color:#8b92a5'>{lbl_}</span>"
                        f"<span style='color:#fff;font-weight:600'>{fmt(v_,1) if v_ else '—'}</span></div>",
                        unsafe_allow_html=True)
        with ss3:
            st.markdown("**Latest Balance Sheet**")
            if bs_rows and yr_cols:
                for lbl_,key_ in [("Total Assets","Total Assets"),("Equity","Equity Capital"),
                                   ("Borrowings","Borrowings"),("Cash","Cash Equivalents")]:
                    s_ = get_series(bs_rows,key_,[yr_cols[-1]])
                    v_ = s_.get(yr_cols[-1])
                    st.markdown(
                        f"<div style='display:flex;justify-content:space-between;"
                        f"padding:5px 0;border-bottom:1px solid #2d3147;font-size:12px'>"
                        f"<span style='color:#8b92a5'>{lbl_}</span>"
                        f"<span style='color:#fff;font-weight:600'>₹{fmt(v_,0) if v_ else '—'} Cr</span></div>",
                        unsafe_allow_html=True)

# ─── Footer ──────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    f"<div style='text-align:center;color:#333;font-size:11px;padding:8px'>"
    f"Equity Research Terminal v4 · Built locally · Data: Screener.in + yfinance + Google News · "
    f"Not financial advice · {datetime.date.today()}"
    f"</div>", unsafe_allow_html=True)
